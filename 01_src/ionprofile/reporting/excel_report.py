"""
excel_report.py - Formatted Excel output
==========================================
Saves profiling results as a formatted Excel workbook with:
    - Color-coded charge columns (red=negative, white=neutral, blue=positive)
    - Auto-fitted column widths
    - Summary sheet with statistics

Project: ionprofile
"""

import logging
from pathlib import Path
from typing import List, Union

import pandas as pd

logger = logging.getLogger(__name__)

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def save_excel(
        df: pd.DataFrame,
        output_path: Union[str, Path],
        ph_values: List[float] = None,
) -> str:
    """
    Save ionization profiling results as formatted Excel.

    Args:
        df:          Results DataFrame.
        output_path: Output file path (.xlsx).
        ph_values:   List of pH values (for charge column formatting).

    Returns:
        Absolute path to saved file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not OPENPYXL_AVAILABLE:
        df.to_excel(output_path, index=False, engine="openpyxl")
        logger.info(f"Saved Excel (plain): {output_path}")
        return str(output_path.resolve())

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # --- Data sheet ---
        df.to_excel(writer, sheet_name="Ionization", index=False)
        ws = writer.sheets["Ionization"]

        # Styles
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50",
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            bottom=Side(style="thin", color="D5D8DC")
        )

        fill_negative = PatternFill(start_color="FADBD8",
                                    end_color="FADBD8", fill_type="solid")
        fill_neutral = PatternFill(start_color="FFFFFF",
                                   end_color="FFFFFF", fill_type="solid")
        fill_positive = PatternFill(start_color="D4E6F1",
                                    end_color="D4E6F1", fill_type="solid")
        fill_na = PatternFill(start_color="F2F3F4",
                              end_color="F2F3F4", fill_type="solid")

        # Format header row
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Identify charge columns
        charge_cols_idx = set()
        if ph_values:
            for ph in ph_values:
                col_name = f"Q_pH{int(round(ph * 10))}"
                if col_name in df.columns:
                    charge_cols_idx.add(
                        list(df.columns).index(col_name) + 1
                    )

        # Format data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                if col_idx in charge_cols_idx:
                    cell.alignment = Alignment(horizontal="center")
                    val = cell.value
                    if val is None or val == "":
                        cell.fill = fill_na
                    elif isinstance(val, (int, float)):
                        if val < 0:
                            cell.fill = fill_negative
                        elif val > 0:
                            cell.fill = fill_positive
                        else:
                            cell.fill = fill_neutral

        # Auto-fit column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max()
                if len(df) > 0 else 0
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                min(max_len + 3, 40)
            )

        # Freeze header row
        ws.freeze_panes = "A2"

        # --- Summary sheet ---
        _write_summary_sheet(writer, df, ph_values)

    logger.info(f"Saved Excel: {output_path} ({len(df)} rows)")
    return str(output_path.resolve())


def _write_summary_sheet(
        writer: pd.ExcelWriter,
        df: pd.DataFrame,
        ph_values: List[float],
) -> None:
    """Write a summary statistics sheet."""
    summary_data = [{"Metric": "Total Molecules", "Value": len(df)}]

    if ph_values:
        summary_data.append({"Metric": "pH Range",
                             "Value": f"{ph_values[0]} -> {ph_values[-1]}"})
        summary_data.append({"Metric": "pH Steps",
                             "Value": len(ph_values)})

        # Distribution at max pH
        col_max = f"Q_pH{int(round(ph_values[0] * 10))}"
        if col_max in df.columns:
            series = df[col_max]
            summary_data.append({"Metric": "", "Value": ""})
            summary_data.append(
                {"Metric": f"--- At pH {ph_values[0]} ---", "Value": ""})
            summary_data.append(
                {"Metric": "Negative", "Value": int((series < 0).sum())})
            summary_data.append(
                {"Metric": "Neutral", "Value": int((series == 0).sum())})
            summary_data.append(
                {"Metric": "Positive", "Value": int((series > 0).sum())})
            summary_data.append(
                {"Metric": "N/A", "Value": int(series.isna().sum())})

        # Distribution at min pH
        col_min = f"Q_pH{int(round(ph_values[-1] * 10))}"
        if col_min in df.columns:
            series = df[col_min]
            summary_data.append({"Metric": "", "Value": ""})
            summary_data.append(
                {"Metric": f"--- At pH {ph_values[-1]} ---", "Value": ""})
            summary_data.append(
                {"Metric": "Negative", "Value": int((series < 0).sum())})
            summary_data.append(
                {"Metric": "Neutral", "Value": int((series == 0).sum())})
            summary_data.append(
                {"Metric": "Positive", "Value": int((series > 0).sum())})

    # Transitions
    if "N_Transitions" in df.columns:
        trans = df["N_Transitions"].dropna()
        if len(trans) > 0:
            summary_data.append({"Metric": "", "Value": ""})
            summary_data.append(
                {"Metric": "--- Transitions ---", "Value": ""})
            summary_data.append(
                {"Metric": "No transition",
                 "Value": int((trans == 0).sum())})
            summary_data.append(
                {"Metric": "Single transition",
                 "Value": int((trans == 1).sum())})
            summary_data.append(
                {"Metric": "Multiple transitions",
                 "Value": int((trans > 1).sum())})

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
